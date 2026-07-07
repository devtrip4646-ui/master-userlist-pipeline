import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

daily = json.load(open("/tmp/daily_rows.json"))
retention = json.load(open("/tmp/retention_rows.json"))

wb = Workbook()

ARIAL = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
AVG_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FONT = Font(name=ARIAL, size=14, bold=True, color="1F4E78")
HEADER_FONT = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name=ARIAL, size=10)
AVG_FONT = Font(name=ARIAL, size=10, bold=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

def style_body_row(ws, row, ncols, start_col=1, money_cols=(), bold=False, fill=None):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = AVG_FONT if bold else BODY_FONT
        cell.border = BORDER
        cell.alignment = CENTER
        if fill:
            cell.fill = fill
        if c in money_cols:
            cell.number_format = "#,##0"

# ---------------- Sheet 1: Daily New vs Old User Analysis ----------------
ws = wb.active
ws.title = "Daily New vs Old Analysis"

ws.merge_cells("A1:K1")
ws["A1"] = "Weekly New vs Old User Deposit & Withdrawal Analysis (30 June - 6 July 2026)"
ws["A1"].font = TITLE_FONT

ws.merge_cells("A2:K2")
ws["A2"] = "Old users = repeat depositors that day (excludes users making their first-ever deposit that day). New users = users whose first-ever deposit landed on that day."
ws["A2"].font = Font(name=ARIAL, size=9, italic=True, color="595959")

headers = [
    "Date",
    "Old Users Count",
    "Avg Deposit of Old Users",
    "New Users Count",
    "Avg Deposit of New Users",
    "Old Users Count (Withdrew)",
    "Avg Withdraw Amount - Old Users",
    "New Users Count (Withdrew)",
    "Avg Withdraw Amount - New Users",
    "Total Deposit (Day)",
    "Total Depositor Count (Day)",
]
HEADER_ROW = 4
for i, h in enumerate(headers, start=1):
    ws.cell(row=HEADER_ROW, column=i, value=h)
style_header(ws, HEADER_ROW, len(headers))

money_cols = {3, 5, 7, 9, 10}
r = HEADER_ROW + 1
for row in daily:
    old_avg_dep = row["old_users_deposit_total"] / row["old_users_count"] if row["old_users_count"] else 0
    new_avg_dep = row["new_users_deposit_total"] / row["new_users_count"] if row["new_users_count"] else 0
    old_avg_wd = row["old_users_withdraw_total"] / row["old_users_withdraw_count"] if row["old_users_withdraw_count"] else 0
    new_avg_wd = row["new_users_withdraw_total"] / row["new_users_withdraw_count"] if row["new_users_withdraw_count"] else 0
    values = [
        row["date"],
        row["old_users_count"],
        round(old_avg_dep, 0),
        row["new_users_count"],
        round(new_avg_dep, 0),
        row["old_users_withdraw_count"],
        round(old_avg_wd, 0),
        row["new_users_withdraw_count"],
        round(new_avg_wd, 0),
        row["total_deposit"],
        row["total_depositor_count"],
    ]
    for i, v in enumerate(values, start=1):
        ws.cell(row=r, column=i, value=v)
    style_body_row(ws, r, len(headers), money_cols=money_cols)
    r += 1

n_days = len(daily)
first_data_row = HEADER_ROW + 1
last_data_row = r - 1

avg_row = r
ws.cell(row=avg_row, column=1, value="7-DAY AVERAGE")
avg_col_letters = [get_column_letter(c) for c in range(2, len(headers) + 1)]
for c in range(2, len(headers) + 1):
    col_letter = get_column_letter(c)
    ws.cell(row=avg_row, column=c,
            value=f"=AVERAGE({col_letter}{first_data_row}:{col_letter}{last_data_row})")
style_body_row(ws, avg_row, len(headers), money_cols=money_cols, bold=True, fill=AVG_FILL)
for c in range(3, len(headers) + 1):
    if c in money_cols or True:
        ws.cell(row=avg_row, column=c).number_format = "#,##0"

col_widths = [12, 16, 22, 16, 22, 22, 26, 22, 26, 20, 22]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "B5"

# ---------------- Sheet 2: New User 3-Day Retention ----------------
ws2 = wb.create_sheet("New User 3-Day Retention")

ws2.merge_cells("A1:H1")
ws2["A1"] = "New User 3-Day Return-Deposit Retention (30 June - 6 July 2026)"
ws2["A1"].font = TITLE_FONT

ws2.merge_cells("A2:H2")
ws2["A2"] = "Retention = % of that day's new users who made a further deposit within the following 3 days, split by whether they withdrew (any status) during that window."
ws2["A2"].font = Font(name=ARIAL, size=9, italic=True, color="595959")

headers2 = [
    "Date",
    "New Users (Day)",
    "Withdrew - Count",
    "Withdrew - Returned (Redeposited)",
    "Withdrew - 3-Day Retention %",
    "Never Withdrew - Count",
    "Never Withdrew - Returned (Redeposited)",
    "Never Withdrew - 3-Day Retention %",
]
HEADER_ROW2 = 4
for i, h in enumerate(headers2, start=1):
    ws2.cell(row=HEADER_ROW2, column=i, value=h)
style_header(ws2, HEADER_ROW2, len(headers2))

r2 = HEADER_ROW2 + 1
for row in retention:
    values = [
        row["date"],
        row["new_users"],
        row["withdrew_group_count"],
        row["withdrew_group_returned"],
        row["withdrew_group_retention_pct"],
        row["never_withdrew_group_count"],
        row["never_withdrew_group_returned"],
        row["never_withdrew_group_retention_pct"],
    ]
    for i, v in enumerate(values, start=1):
        ws2.cell(row=r2, column=i, value=v)
    for c in (5, 8):
        ws2.cell(row=r2, column=c).number_format = "0.00\"%\""
    style_body_row(ws2, r2, len(headers2))
    r2 += 1

first2 = HEADER_ROW2 + 1
last2 = r2 - 1
avg_row2 = r2
ws2.cell(row=avg_row2, column=1, value="7-DAY AVERAGE")
for c in range(2, len(headers2) + 1):
    col_letter = get_column_letter(c)
    ws2.cell(row=avg_row2, column=c,
             value=f"=AVERAGE({col_letter}{first2}:{col_letter}{last2})")
style_body_row(ws2, avg_row2, len(headers2), bold=True, fill=AVG_FILL)
for c in (5, 8):
    ws2.cell(row=avg_row2, column=c).number_format = "0.00\"%\""

col_widths2 = [12, 16, 16, 26, 22, 18, 28, 24]
for i, w in enumerate(col_widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "B5"

wb.save("/Users/devtr/Downloads/weekly_new_old_user_analysis.xlsx")
print("saved")
